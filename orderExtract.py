#! python3 - Event Order # Regex and PDF extraction
#	UPDATED - 2.1.2017

import os
import PyPDF2
import re
import openpyxl
import warnings

def orderEx():

    #---#---#--- SYS ARGV FOR COMMAND LINE INTERFACE ---#---#---#

    ''' Handles wrong file names being entered'''

    status = False

    while status == False:
        pdf = 'PATH TO EVENT CHECK PDF' + input('What month?: ') + '\\'+ input('Enter PDF file: ') + '.pdf'
        try:
            file = open(pdf, 'rb')
            warnings.simplefilter("ignore") 
            workbook = openpyxl.load_workbook('PATH TO GRAUTUITY TEMPLATE - EXCEL')# template file
            reader = PyPDF2.PdfFileReader(file)
        except FileNotFoundError:
            print('File not found, try again.')
            continue
        status = True
        

    totalGrat = 0   
    noGrat = 0
    count = 0
    colNum = 5
    sheet = workbook.get_sheet_by_name('GRATS')

    ''' Contains the Date, Order #, and Gratuity Regex
        and tallies all of the grauity for the period.'''

    while count < reader.numPages:
            
            try:
                    eventCheck = reader.getPage(count).extractText()
                    
                    orderRegex = re.compile(r'Order #:(\d+)')
                    orderNum = orderRegex.findall(eventCheck)

                    dateRegex = re.compile(r'\w+\s\d+\D\s\d+')
                    date = dateRegex.findall(eventCheck)

                    gratRegex = re.compile(r'22\%\$(\d\D?\d+?\D\d+)')
                    gratuity = gratRegex.findall(eventCheck)
                    
                    totalGrat += float(gratuity[0].replace(',', ''))
                    
                    print('\n' + date[0])
                    print('Event Order #: ' + orderNum[0])
                    print('Banquet Gratuity: $' + gratuity[0])
                    print('Page ' + str(count) + ' of ' + str(reader.numPages))
                    
            except IndexError:
                    print('\nNo Gratuity Found on page #: ' + str(count))
                    noGrat += 1
                    pass
                    
            count += 1
            
            #-----#-----#----- EXCEL DATA -----#-----#-----#
            try:
                    sheet.cell(row=1, column=colNum).value = date[0]
                    sheet.cell(row=2, column=colNum).value = int(orderNum[0])
                    sheet.cell(row=4, column=colNum).value = float(gratuity[0].replace(',', ''))
            except IndexError:
                    pass
                    
            colNum+= 1


    print('\n' + ('#-----#-----')*5)
    print('\n' + 'Total Gratuity: '+'$'+str(totalGrat))
    print(str(count) + ' total pages scanned.')
    print(str(noGrat) + ' pages without gratuity.')

    print('\n@******@******@******@*******@\n')
    
    #TODO Need to Put in a check to see if the month directory is made, if it isn't then needs to create New Folder for that month

    workbook.save('PATH TO SAVE NEW GRATUITY SHEET - EXCEL'+ input('What month?: ') + '\\' + input('Save spreadsheet as: ')+'.xlsx')
     
    
orderEx() 
        
        
        
